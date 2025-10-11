#!/usr/bin/env python3
import openpyxl
import psycopg2
import os
import sys
from collections import defaultdict

# Connessione al database
conn = psycopg2.connect(
    host="localhost",
    database="assessment_db",
    user="assessment_user",
    password="bAX2lyThZOSk0nmwwT6R7mva1bTHN+zJq2JugLHDykg="
)
cur = conn.cursor()

# Leggi Excel dal path corretto
excel_path = 'frontend/public/file di controllo.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Mappa categorie -> colonne
category_cols = {
    'Governance': (2, 6),  # B-F (5 dimensioni)
    'Monitoring & Control': (8, 11),  # H-K (4 dimensioni)
    'Technology': (13, 15),  # M-O (3 dimensioni)
    'Organization': (17, 18)  # Q-R (2 dimensioni)
}

# Leggi headers (riga 2)
headers = {}
for cat, (start, end) in category_cols.items():
    headers[cat] = [ws.cell(2, col).value for col in range(start, end + 1) if ws.cell(2, col).value and 'note' not in str(ws.cell(2, col).value).lower()]

print("=== HEADERS ===")
for cat, dims in headers.items():
    print(f"{cat}: {len(dims)} dimensioni")
    for d in dims:
        print(f"  - {d}")

# Leggi dati Excel
expected_data = []
current_process = None

for row_idx in range(3, ws.max_row + 1):
    activity_name = ws.cell(row_idx, 1).value
    
    if not activity_name:
        continue
    
    # Check se è un nome di processo (tutte le altre celle vuote)
    is_process = all(ws.cell(row_idx, col).value is None for col in range(2, 20))
    
    if is_process:
        current_process = activity_name
        print(f"\n--- PROCESSO: {current_process} ---")
        continue
    
    if not current_process:
        continue
    
    # Leggi scores per ogni categoria
    for category, (start_col, end_col) in category_cols.items():
        for dim_idx, col in enumerate(range(start_col, end_col + 1)):
            header = ws.cell(2, col).value
            if not header or 'note' in str(header).lower():
                continue
            
            score = ws.cell(row_idx, col).value
            if score is not None:
                expected_data.append({
                    'process': current_process,
                    'activity': activity_name,
                    'category': category,
                    'dimension': header,
                    'score': score
                })

print(f"\n=== DATI EXCEL ===")
print(f"Totale righe attese: {len(expected_data)}")

# Query database per trovare session_id recenti
cur.execute("SELECT id, azienda_nome, creato_il FROM assessment_session ORDER BY creato_il DESC LIMIT 5")
sessions = cur.fetchall()
print("\n=== SESSIONI RECENTI ===")
for sess in sessions:
    print(f"ID: {sess[0]} | Azienda: {sess[1]} | Data: {sess[2]}")

if len(sys.argv) > 1:
    session_id = sys.argv[1]
else:
    print("\nUsage: python3 verify_assessment_data.py <session_id>")
    print("Oppure digita il session_id:")
    session_id = input().strip()

# Verifica dati
cur.execute("""
    SELECT process, activity, category, dimension, score, is_not_applicable
    FROM assessment_result 
    WHERE session_id = %s
    ORDER BY process, activity, category, dimension
""", (session_id,))

db_data = cur.fetchall()
print(f"\n=== DATI DATABASE ===")
print(f"Totale righe nel DB: {len(db_data)}")

# Confronta
print("\n=== CONFRONTO ===")
discrepancies = []

for expected in expected_data:
    # Cerca nel DB
    found = False
    for db_row in db_data:
        if (db_row[0] == expected['process'] and 
            db_row[1] == expected['activity'] and
            db_row[2] == expected['category'] and
            db_row[3] == expected['dimension']):
            found = True
            if db_row[4] != expected['score']:
                discrepancies.append({
                    'type': 'VALORE DIVERSO',
                    'process': expected['process'],
                    'activity': expected['activity'],
                    'category': expected['category'],
                    'dimension': expected['dimension'],
                    'expected': expected['score'],
                    'actual': db_row[4]
                })
            break
    
    if not found:
        discrepancies.append({
            'type': 'MANCANTE NEL DB',
            'process': expected['process'],
            'activity': expected['activity'],
            'category': expected['category'],
            'dimension': expected['dimension'],
            'expected': expected['score'],
            'actual': None
        })

if discrepancies:
    print(f"\n❌ Trovate {len(discrepancies)} discrepanze:")
    for d in discrepancies[:20]:  # Mostra prime 20
        print(f"\n{d['type']}")
        print(f"  Processo: {d['process']}")
        print(f"  Attività: {d['activity']}")
        print(f"  Categoria: {d['category']}")
        print(f"  Dimensione: {d['dimension'][:50]}...")
        print(f"  Atteso: {d['expected']}, Trovato: {d['actual']}")
else:
    print("\n✅ Tutti i dati corrispondono!")

cur.close()
conn.close()
wb.close()
