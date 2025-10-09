#!/usr/bin/env python3
"""
Script per importare dati da Excel compilato nel database
"""
import openpyxl
import sys
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import uuid
from datetime import datetime

# Importa i modelli
sys.path.insert(0, '/var/www/assessment_ai')
from app import models
from app.database import SQLALCHEMY_DATABASE_URL

def import_excel_to_db(excel_path, company_name, sector, size, referente, email):
    """Importa dati da Excel nel database"""
    
    # Connessione DB
    engine = create_engine(SQLALCHEMY_DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()
    
    try:
        # 1. Crea sessione assessment
        session_id = uuid.uuid4()
        new_session = models.AssessmentSession(
            id=session_id,
            azienda_nome=company_name,
            settore=sector,
            dimensione=size,
            referente=referente,
            email=email,
            model_name="Casoinfinal",
            creato_il=datetime.utcnow()
        )
        db.add(new_session)
        db.flush()
        
        print(f"✅ Sessione creata: {session_id}")
        
        # 2. Leggi Excel - sheet CASOIN
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['CASOIN']  # Sheet specifico
        
        # 3. Leggi headers dalla riga 3
        headers = {}
        for cell in ws[3]:
            if cell.value:
                headers[cell.column] = cell.value
        
        print(f"📋 Headers: {len(headers)} colonne")
        
        # 4. Mappa categorie alle colonne (score + nota)
        category_map = {
            # Governance: colonne B-F (score) + G (note)
            "Governance": {
                "scores": list(range(2, 7)),
                "note_col": 7  # Colonna G
            },
            # Monitoring & Control: colonne I-L (score) + M (note)
            "Monitoring & Control": {
                "scores": list(range(9, 13)),
                "note_col": 13  # Colonna M
            },
            # Technology: colonne O-Q (score) + R (note)
            "Technology": {
                "scores": list(range(15, 18)),
                "note_col": 18  # Colonna R
            },
            # Organization: colonne T-U (score) + V (note)
            "Organization": {
                "scores": list(range(20, 22)),
                "note_col": 22  # Colonna V
            }
        }
        
        # 5. Processa righe (dalla riga 4)
        results_count = 0
        current_process = None
        
        for row_idx in range(4, ws.max_row + 1):
            row = ws[row_idx]
            name = row[0].value  # Colonna A
            first_score = row[1].value  # Colonna B
            
            if not name:
                continue
            
            # Se B è None o formula, è un PROCESSO
            is_process_header = (first_score is None or 
                               (isinstance(first_score, str) and first_score.startswith('=')))
            
            if is_process_header:
                current_process = name
                print(f"\n📦 PROCESSO: {current_process}")
                continue
            
            # È un'ATTIVITÀ
            activity = name
            if not current_process:
                print(f"⚠️  Attività senza processo: {activity}")
                continue
            
            print(f"  └─ {activity}")
            
            # Processa ogni categoria
            for category, col_config in category_map.items():
                # Leggi la nota specifica per questa categoria
                note_col_idx = col_config["note_col"]
                note_val = row[note_col_idx - 1].value if note_col_idx <= len(row) else None
                note_text = str(note_val).strip() if note_val else None
                
                # Processa gli score di questa categoria
                for col_idx in col_config["scores"]:
                    score_val = row[col_idx - 1].value  # openpyxl è 1-indexed
                    dimension = headers.get(col_idx, f"Unknown_{col_idx}")
                    
                    # Salta colonne vuote o con nomi speciali
                    if not dimension or 'note' in dimension.lower() or 'total' in dimension.lower():
                        continue
                    
                    # Converti score
                    if isinstance(score_val, (int, float)):
                        score = int(score_val)
                        if 0 <= score <= 5:
                            result = models.AssessmentResult(
                                session_id=session_id,
                                process=current_process,
                                activity=activity,
                                category=category,
                                dimension=dimension,
                                score=score,
                                note=note_text,  # ✅ Nota specifica per categoria
                                is_not_applicable=(score == 0)
                            )
                            db.add(result)
                            results_count += 1
                            
                            if note_text:
                                print(f"    📝 {category}: {dimension} = {score} (nota: {note_text[:50]}...)")
        
        db.commit()
        print(f"\n✅ Importati {results_count} risultati")
        print(f"🔗 Session ID: {session_id}")
        return str(session_id)
        
    except Exception as e:
        db.rollback()
        print(f"❌ Errore: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        db.close()

if __name__ == "__main__":
    print("🚀 IMPORT EXCEL TO DATABASE")
    print("="*80)
    
    session_id = import_excel_to_db(
        excel_path="frontend/public/20250729I40Assessmentmicro.xlsx",
        company_name="Casoin",
        sector="Alimentare",
        size="Micro (1-9 dipendenti)",
        referente="Marco",
        email="marco@casoin.it"
    )
    
    if session_id:
        print(f"\n🎉 IMPORT COMPLETATO!")
        print(f"🔗 Visualizza risultati su:")
        print(f"   https://assessment.noscite.it/results/{session_id}")
