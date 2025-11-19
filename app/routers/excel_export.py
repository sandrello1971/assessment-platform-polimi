from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, Reference
import json
import io
from pathlib import Path

router = APIRouter()

@router.get("/export-model-excel/{model_name}")
async def export_model_to_excel(model_name: str):
    """
    Esporta un modello in formato Excel completo con:
    - Sheet per ogni processo (compilazione dati)
    - Sheet "Riepilogo Risultati" con formule automatiche
    - Sheet "Gap Analysis" 
    - Calcoli automatici che replicano la pagina online
    """
    
    # Carica il modello JSON
    model_path = Path(f"frontend/public/{model_name}.json")
    if not model_path.exists():
        raise HTTPException(status_code=404, detail=f"Modello {model_name} non trovato")
    
    with open(model_path, 'r', encoding='utf-8') as f:
        model_data = json.load(f)
    
    # Crea il workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Stili comuni
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    
    domain_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    domain_header_font = Font(bold=True, color="FFFFFF", size=12)
    
    question_header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    question_header_font = Font(bold=True, size=10)
    
    activity_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    media_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    strength_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
    weakness_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rosso
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Estrai categorie e domande
    categories = {
        'Governance': [],
        'Monitoring & Control': [],
        'Technology': [],
        'Organization': []
    }
    
    if model_data and len(model_data) > 0:
        first_process = model_data[0]
        if 'activities' in first_process and len(first_process['activities']) > 0:
            first_activity = first_process['activities'][0]
            if 'categories' in first_activity:
                for cat_name, questions in first_activity['categories'].items():
                    if cat_name in categories:
                        categories[cat_name] = list(questions.keys())
    
    # ============================================
    # SHEET 1: RIEPILOGO RISULTATI
    # ============================================
    summary_ws = wb.create_sheet(title="Riepilogo Risultati", index=0)
    
    # Header principale
    summary_ws.merge_cells('A1:G1')
    summary_ws['A1'] = 'RIEPILOGO RISULTATI ASSESSMENT'
    summary_ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    summary_ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    summary_ws.row_dimensions[1].height = 35
    
    current_row = 3
    
    # Sezione: PUNTEGGI PER PROCESSO E DIMENSIONE
    summary_ws[f'A{current_row}'] = 'PUNTEGGI PER PROCESSO E DIMENSIONE'
    summary_ws[f'A{current_row}'].font = Font(bold=True, size=14)
    summary_ws.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Header tabella
    headers = ['Processo', 'Governance', 'Monitoring & Control', 'Technology', 'Organization', 'Media Totale', 'Gap %']
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = question_header_font
        cell.fill = question_header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    summary_ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        summary_ws.column_dimensions[col].width = 15
    
    current_row += 1
    process_data_start_row = current_row
    
    # Mappa: nome processo -> nome sheet (troncato a 31 caratteri)
    process_sheet_map = {}
    
    # Righe per ogni processo con formule che leggono dagli sheet processo
    for process in model_data:
        process_name = process['process']
        sheet_name = process_name[:31]
        process_sheet_map[process_name] = sheet_name
        
        # Nome processo
        summary_ws.cell(row=current_row, column=1).value = process_name
        summary_ws.cell(row=current_row, column=1).border = border
        summary_ws.cell(row=current_row, column=1).fill = activity_fill
        
        # Formule per calcolare medie per dimensione
        # Per ogni dimensione: AVERAGE delle colonne "Media" delle sezioni di quel dominio
        for dim_idx, (domain_name, _) in enumerate(categories.items(), start=2):
            cell = summary_ws.cell(row=current_row, column=dim_idx)
            # Formula placeholder - verrà popolata dopo aver creato gli sheet
            cell.value = 0
            cell.border = border
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Media totale (media delle 4 dimensioni)
        cell = summary_ws.cell(row=current_row, column=6)
        cell.value = f'=AVERAGE(B{current_row}:E{current_row})'
        cell.border = border
        cell.fill = media_fill
        cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        
        # Gap % (placeholder)
        cell = summary_ws.cell(row=current_row, column=7)
        cell.value = f'=(5-F{current_row})/5*100'
        cell.border = border
        cell.number_format = '0.0"%"'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
    
    process_data_end_row = current_row - 1
    current_row += 2
    
    # Sezione: RADAR CHART DATA
    summary_ws[f'A{current_row}'] = 'DATI RADAR CHART - DIMENSIONI'
    summary_ws[f'A{current_row}'].font = Font(bold=True, size=14)
    current_row += 1
    
    # Media per dimensione (su tutti i processi)
    summary_ws.cell(row=current_row, column=1).value = 'Dimensione'
    summary_ws.cell(row=current_row, column=2).value = 'Punteggio Medio'
    for col in range(1, 3):
        summary_ws.cell(row=current_row, column=col).font = question_header_font
        summary_ws.cell(row=current_row, column=col).fill = question_header_fill
        summary_ws.cell(row=current_row, column=col).border = border
    current_row += 1
    
    radar_start_row = current_row
    for dim_idx, domain_name in enumerate(categories.keys()):
        col_letter = get_column_letter(2 + dim_idx)
        summary_ws.cell(row=current_row, column=1).value = domain_name
        summary_ws.cell(row=current_row, column=2).value = f'=AVERAGE({col_letter}{process_data_start_row}:{col_letter}{process_data_end_row})'
        summary_ws.cell(row=current_row, column=2).number_format = '0.00'
        for col in range(1, 3):
            summary_ws.cell(row=current_row, column=col).border = border
        current_row += 1
    
    current_row += 2
    
    # Sezione: TOP 3 PROCESSI (Strengths)
    summary_ws[f'A{current_row}'] = 'TOP 3 PROCESSI (Punti di Forza)'
    summary_ws[f'A{current_row}'].font = Font(bold=True, size=12, color="008000")
    summary_ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    summary_ws[f'A{current_row}'] = 'Ordina la tabella sopra per "Media Totale" in ordine decrescente per identificare i top 3'
    current_row += 2
    
    # Sezione: BOTTOM 3 PROCESSI (Weaknesses)
    summary_ws[f'A{current_row}'] = 'BOTTOM 3 PROCESSI (Aree di Miglioramento)'
    summary_ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FF0000")
    summary_ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    summary_ws[f'A{current_row}'] = 'Ordina la tabella sopra per "Media Totale" in ordine crescente per identificare i bottom 3'
    
    # ============================================
    # SHEET 2: GAP ANALYSIS
    # ============================================
    gap_ws = wb.create_sheet(title="Gap Analysis", index=1)
    
    gap_ws.merge_cells('A1:E1')
    gap_ws['A1'] = 'GAP ANALYSIS - STRENGTHS & WEAKNESSES'
    gap_ws['A1'].font = header_font
    gap_ws['A1'].fill = header_fill
    gap_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    gap_ws.row_dimensions[1].height = 30
    
    gap_row = 3
    
    for process in model_data:
        process_name = process['process']
        
        # Header processo
        gap_ws[f'A{gap_row}'] = process_name
        gap_ws[f'A{gap_row}'].font = Font(bold=True, size=14)
        gap_ws.merge_cells(f'A{gap_row}:E{gap_row}')
        gap_ws[f'A{gap_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        gap_row += 1
        
        # Header tabella
        gap_ws[f'A{gap_row}'] = 'Dominio'
        gap_ws[f'B{gap_row}'] = 'Punteggio'
        gap_ws[f'C{gap_row}'] = 'Strengths (≥ 3.0)'
        gap_ws[f'D{gap_row}'] = 'Weaknesses (< 2.0)'
        for col in range(1, 5):
            gap_ws.cell(row=gap_row, column=col).font = question_header_font
            gap_ws.cell(row=gap_row, column=col).fill = question_header_fill
            gap_ws.cell(row=gap_row, column=col).border = border
        gap_row += 1
        
        # Righe domini
        for domain_name in categories.keys():
            gap_ws.cell(row=gap_row, column=1).value = domain_name
            gap_ws.cell(row=gap_row, column=2).value = 0  # Verrà linkato
            gap_ws.cell(row=gap_row, column=3).value = ''  # Inserire manualmente
            gap_ws.cell(row=gap_row, column=4).value = ''  # Inserire manualmente
            
            for col in range(1, 5):
                gap_ws.cell(row=gap_row, column=col).border = border
            
            # Colora in base al punteggio
            score_cell = gap_ws.cell(row=gap_row, column=2)
            score_cell.number_format = '0.00'
            
            gap_row += 1
        
        gap_row += 1
    
    gap_ws.column_dimensions['A'].width = 25
    gap_ws.column_dimensions['B'].width = 12
    gap_ws.column_dimensions['C'].width = 40
    gap_ws.column_dimensions['D'].width = 40
    
    # ============================================
    # CREA SHEET PER OGNI PROCESSO
    # ============================================
    for proc_idx, process in enumerate(model_data):
        process_name = process['process']
        activities = process['activities']
        
        sheet_name = process_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        max_questions = max(len(q) for q in categories.values() if q)
        total_cols = 1 + max_questions + 1 + 1
        
        # HEADER PROCESSO
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        ws['A1'] = process_name
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        row_idx = 2
        
        # Dizionario per tenere traccia delle righe di ogni dominio per le formule
        domain_rows = {}
        
        # Per ogni dominio
        for domain_name, questions in categories.items():
            if not questions:
                continue
            
            domain_start_row = row_idx
            
            # HEADER DOMINIO
            ws.merge_cells(f'A{row_idx}:{get_column_letter(total_cols)}{row_idx}')
            ws[f'A{row_idx}'] = domain_name
            ws[f'A{row_idx}'].font = domain_header_font
            ws[f'A{row_idx}'].fill = domain_header_fill
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row_idx].height = 25
            row_idx += 1
            
            # HEADER COLONNE
            ws[f'A{row_idx}'] = 'Attività'
            ws[f'A{row_idx}'].font = question_header_font
            ws[f'A{row_idx}'].fill = question_header_fill
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[f'A{row_idx}'].border = border
            ws.column_dimensions['A'].width = 35
            
            col_idx = 2
            for question in questions:
                ws[f'{get_column_letter(col_idx)}{row_idx}'] = question
                ws[f'{get_column_letter(col_idx)}{row_idx}'].font = question_header_font
                ws[f'{get_column_letter(col_idx)}{row_idx}'].fill = question_header_fill
                ws[f'{get_column_letter(col_idx)}{row_idx}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[f'{get_column_letter(col_idx)}{row_idx}'].border = border
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
                col_idx += 1
            
            media_col = col_idx
            ws[f'{get_column_letter(media_col)}{row_idx}'] = 'Media'
            ws[f'{get_column_letter(media_col)}{row_idx}'].font = question_header_font
            ws[f'{get_column_letter(media_col)}{row_idx}'].fill = question_header_fill
            ws[f'{get_column_letter(media_col)}{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{get_column_letter(media_col)}{row_idx}'].border = border
            ws.column_dimensions[get_column_letter(media_col)].width = 10
            
            note_col = media_col + 1
            ws[f'{get_column_letter(note_col)}{row_idx}'] = 'Note'
            ws[f'{get_column_letter(note_col)}{row_idx}'].font = question_header_font
            ws[f'{get_column_letter(note_col)}{row_idx}'].fill = question_header_fill
            ws[f'{get_column_letter(note_col)}{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{get_column_letter(note_col)}{row_idx}'].border = border
            ws.column_dimensions[get_column_letter(note_col)].width = 40
            
            ws.row_dimensions[row_idx].height = 50
            row_idx += 1
            
            activities_start_row = row_idx
            
            # RIGHE ATTIVITÀ
            for activity in activities:
                ws[f'A{row_idx}'] = activity['name']
                ws[f'A{row_idx}'].border = border
                ws[f'A{row_idx}'].fill = activity_fill
                ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                question_cols = []
                for q_idx in range(len(questions)):
                    col = get_column_letter(2 + q_idx)
                    ws[f'{col}{row_idx}'].border = border
                    ws[f'{col}{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                    question_cols.append(col)
                
                # Formula Media: AVERAGE ignora automaticamente testo e celle vuote
                range_start = question_cols[0]
                range_end = question_cols[-1]
                media_cell = ws[f'{get_column_letter(media_col)}{row_idx}']
                media_cell.value = f'=AVERAGE({range_start}{row_idx}:{range_end}{row_idx})'
                media_cell.border = border
                media_cell.fill = media_fill
                media_cell.alignment = Alignment(horizontal='center', vertical='center')
                media_cell.number_format = '0.00'
                
                ws[f'{get_column_letter(note_col)}{row_idx}'].border = border
                ws[f'{get_column_letter(note_col)}{row_idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1
            
            activities_end_row = row_idx - 1
            
            # Salva info per formule Riepilogo
            domain_rows[domain_name] = {
                'sheet': sheet_name,
                'media_col': get_column_letter(media_col),
                'start_row': activities_start_row,
                'end_row': activities_end_row
            }
            
            row_idx += 1
        
        # Ora aggiorna le formule nel Riepilogo per questo processo
        # Trova la riga di questo processo nel Riepilogo
        summary_row = process_data_start_row + proc_idx
        
        for dim_idx, domain_name in enumerate(categories.keys(), start=2):
            if domain_name in domain_rows:
                info = domain_rows[domain_name]
                formula = f"=AVERAGE('{info['sheet']}'!{info['media_col']}{info['start_row']}:{info['media_col']}{info['end_row']})"
                summary_ws.cell(row=summary_row, column=dim_idx).value = formula
    
    # Salva
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"{model_name}_assessment.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
